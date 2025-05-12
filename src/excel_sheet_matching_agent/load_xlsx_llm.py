from dotenv import load_dotenv
from pydantic import BaseModel, Field
from typing import List, Union
from langchain_core.language_models.chat_models import BaseChatModel
from langfuse.callback import CallbackHandler
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter # 列番号を列名に変換するために使用

import logging

from .models import ExcelCellInputData
from .prompts import extract_inputs_prompt

load_dotenv()
logger = logging.getLogger(__name__)


def sheet2str(sheet: Worksheet) -> str:
    """
    openpyxlのシートをHTMLテーブル文字列に変換する関数
    """
    rows = []
    for row_idx, row in enumerate(sheet.iter_rows()):
        excel_row_num = row_idx + 1 # Excelの行番号は1から始まる
        tr = "<tr>"
        for col_idx, cell in enumerate(row):
            excel_col_letter = get_column_letter(col_idx + 1)
            cell_value = cell.value

            if cell_value is None:
                cell_value = ""

            if isinstance(cell_value, (int, float)):
                tr += f'<td data-row="{excel_row_num}" data-col="{excel_col_letter}" class="input">{cell_value}</td>'
            elif isinstance(cell_value, str) and cell.data_type == "f":
                tr += f'<td data-row="{excel_row_num}" data-col="{excel_col_letter}" class="formula">{cell_value}</td>'
            else:
                tr += f'<td data-row="{excel_row_num}" data-col="{excel_col_letter}">{cell_value}</td>'
        tr += "</tr>"
        rows.append(tr)

    return "<table>\n" + "\n".join(rows) + "\n</table>"


class _ExcelCellInputData(BaseModel):
    cell: str = Field(..., description="Excel sheet cell name, like A1, B2")
    value: Union[int, float] = Field(..., description="Integer or float value from the cell")
    metadata: List[str] = Field(..., description="Contextual information or labels associated with the value, inferred from surrounding cells (e.g., headers, row labels). Prioritize extracting units if they are present near the cell.")


class _ExcelCellInputDatas(BaseModel):
    inputs: List[_ExcelCellInputData] = Field(default_factory=list)


def extract_inputs(llm: BaseChatModel, sheet: Worksheet) -> List[ExcelCellInputData]:
    html_output = sheet2str(sheet)
    chain = extract_inputs_prompt | llm.with_structured_output(_ExcelCellInputDatas)
    result: _ExcelCellInputDatas = chain.invoke({"html_table": html_output}, config={"callbacks": [CallbackHandler()]})  # type: ignore

    inputs = []
    for data in result.inputs:
        logger.debug(data.model_dump_json(indent=2, exclude_none=True))
        inputs.append(ExcelCellInputData(
            sheet=sheet.title,
            cell=data.cell,
            value=data.value,
            metadata=data.metadata
        ))

    return inputs
