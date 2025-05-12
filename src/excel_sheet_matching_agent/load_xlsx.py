"""
Excel Worksheetから入力セル/計算セルを抽出する
"""
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pathlib import Path

from .models import ExcelCellFormulaData, ExcelCellInputData, ExcelSheetData


def extract_cell_info(sheet: Worksheet) -> ExcelSheetData:
    input_data = []
    formula_data = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            if isinstance(cell.value, (int, float)):
                cell_type = "input"
            elif isinstance(cell.value, str) and cell.data_type == "f":
                cell_type = "formula"
            else:
                continue

            # 周囲セルの情報（左右）
            def safe_value(col_offset, row_offset):
                try:
                    ref_cell = sheet.cell(row=cell.row + row_offset, column=cell.column + col_offset)
                    return str(ref_cell.value).strip() if ref_cell and ref_cell.value else ""
                except:
                    return ""

            metadata = [safe_value(-1, 0), safe_value(1, 0)]

            if cell_type == "input":
                input_data.append(ExcelCellInputData(
                    sheet=sheet.title,
                    cell=cell.coordinate,
                    value=cell.value,  # type: ignore
                    metadata=metadata
                ))
            elif cell_type == "formula":
                formula_data.append(ExcelCellFormulaData(
                    sheet=sheet.title,
                    cell=cell.coordinate,
                    value=cell.value, # type: ignore
                    metadata=metadata
                ))

    return ExcelSheetData(input=input_data, formula=formula_data)


def extract_data(sheet_path: Path, sheet_name: str) -> ExcelSheetData:
    wb = load_workbook(sheet_path, data_only=False)
    for sheet in wb.worksheets:
        if sheet.title != sheet_name:
            continue
        return extract_cell_info(sheet)
    raise ValueError(f"sheet not found: {sheet_name}")
