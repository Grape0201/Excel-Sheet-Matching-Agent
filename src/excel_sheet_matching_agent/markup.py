'''
matching結果を元に、Excel印刷PDFと出典PDFをマークアップする（赤文字記号を書き込む）
'''
from azure.ai.documentintelligence.models import AnalyzeResult
import fitz
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from pathlib import Path
from typing import Optional
import json
import logging
import csv

from .models import MatchingResult, ExcelCellInputData


MARK_SYMBOLS = "あいうえお"

logger = logging.getLogger(__name__)

def load_prebuild_layout_result(json_path: Path) -> AnalyzeResult:
    with json_path.open(encoding='utf-8') as f:
        return AnalyzeResult(json.load(f))

def markup_source_pdf(
    source_pdf_path: Path,
    output_pdf_path: Path,
    results: list[MatchingResult],
    prebuilt_layout_result: AnalyzeResult,
    symbols: str,
) -> list[int]:
    doc = fitz.open(source_pdf_path)

    page_numbers = []
    for res, symbol in zip(results, symbols):
        if not res.match or not res.matched_text:
            page_numbers.append(0)
            continue

        # 1. 行単位で候補行を探す
        candidate = None
        for page in prebuilt_layout_result.pages:
            for line in page.lines or []:
                if res.matched_text in line.content:
                    candidate = (page, line)
                    break
            if candidate:
                break

        if candidate is None:
            logger.warning(f"no line candidate for: {res.cell}")
            page_numbers.append(0)
            continue

        page_obj, line = candidate
        page = doc[page_obj.page_number - 1]
        page_numbers.append(page_obj.page_number)

        # 2. 文字単位ポリゴンの結合
        xs0, ys0, xs1, ys1 = [], [], [], []
        for word in page_obj.words or []:
            if word.content in res.matched_text:
                poly = word.polygon
                xs = poly[0::2]; ys = poly[1::2]  # type: ignore
                xs0.append(min(xs)); ys0.append(min(ys))
                xs1.append(max(xs)); ys1.append(max(ys))

        if xs0:
            # 文字単位矩形取得
            x0, y0 = min(xs0), min(ys0)
            x1, y1 = max(xs1), max(ys1)
        else:
            # フォールバック：行ポリゴン中央
            poly = line.polygon
            px = [(poly[i] + poly[i+2])/2 for i in range(0, 8, 2)]  # type: ignore
            py = [(poly[i+1] + poly[i+3])/2 for i in range(0, 8, 2)]  # type: ignore
            x0, y0 = sum(px)/4, sum(py)/4
            x1, y1 = x0 + 10, y0 + 10  # small marker

        # 3. PDF 座標系への変換（inch→pt）
        # Azure: inch 単位。fitz は pt(1pt=1/72in)
        rect = fitz.Rect(x0*72, y0*72, x1*72, y1*72)
        font_size = rect.height * 0.8

        # 4. マーカー挿入
        page.insert_text(  # type: ignore
            (rect.x0, rect.y0),
            symbol,
            fontsize=font_size,
            fontname="japan",
            color=(1, 0, 0)
        )

    logger.info(f"writing: {output_pdf_path}")
    doc.save(output_pdf_path)
    return page_numbers

def _normalize_num_str(s: str) -> str:
    try:
        f = float(s)
    except (TypeError, ValueError):
        return s.strip()
    if f.is_integer():
        return str(int(f))
    return str(f)

def _find_pdf_bbox_for_cell(page, ws: Worksheet, cell: str) -> Optional[tuple[float, float, float, float]]:
    """
    PDFページ上で target_text と完全一致するテキストスパンを探し、
    その bbox を返す。見つからなければ None。
    """
    raw = ws[cell].value
    if raw is None:
        return None
    target_norm = _normalize_num_str(str(raw))

    words = page.get_text("words")
    for x0, y0, x1, y1, text, *_ in words:
        if _normalize_num_str(text) == target_norm:
            return (x0, y0, x1, y1)
    return None

def markup_excel_pdf(
    excel_path: Path,
    source_pdf_path: Path,
    output_pdf_path: Path,
    inputs: list[ExcelCellInputData],
    matches: list[bool],
    mark_symbols: str
):
    assert len(inputs) == len(matches)
    # Excel読み込み
    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws is not None

    # PDF読み込み
    doc = fitz.open(source_pdf_path)
    page = doc[0]  # TODO p.1

    for inp, match, symbol in zip(inputs, matches, mark_symbols):
        if not match:
            continue
        cell = inp.cell
        bbox = _find_pdf_bbox_for_cell(page, ws, cell)
        if not bbox:
            logger.warning(f"{cell} が見つかりません")
            continue
        _, _, x1, y1 = bbox

        # テキスト挿入
        page.insert_text(  # type: ignore
            fitz.Point(x1 + 2, y1 - 2),
            symbol,
            fontsize=12,
            fontname="japan",
            color=(1, 0, 0)
         ) # :contentReference[oaicite:6]{index=6}

    # 保存
    logger.info(f"writing: {output_pdf_path}")
    doc.save(output_pdf_path)

def markup(
        excel_path: Path,
        excel_pdf_path: Path,
        inputs: list[ExcelCellInputData],
        matching_results: list[MatchingResult],
        source_pdfs: list[Path],
        prebuilt_layout_result_jsons: list[Path],
        mark_symbols: str = MARK_SYMBOLS
    ):
    assert len(source_pdfs) == len(prebuilt_layout_result_jsons)
    assert len(inputs) == len(matching_results)
    assert len(MARK_SYMBOLS) > len(inputs)
    assert excel_path.exists()
    assert excel_pdf_path.exists()
    for source_pdf in source_pdfs:
        assert source_pdf.exists()
    for prebuilt_layout_result_json in prebuilt_layout_result_jsons:
        assert prebuilt_layout_result_json.exists()

    # Excel
    matches = [match.match for match in matching_results]
    excel_pdf_markup_path = excel_pdf_path.parent / f"{excel_pdf_path.stem}_markup.pdf"
    markup_excel_pdf(excel_path, excel_pdf_path, excel_pdf_markup_path, inputs, matches, mark_symbols)

    # Source
    source_pages = [("", 0)]*len(inputs)
    for source_pdf, prebuild_layout_json in zip(source_pdfs, prebuilt_layout_result_jsons):
        source_pdf_markup_path = source_pdf.parent / f"{source_pdf.stem}_markup.pdf"
        prebuilt_layout_result = load_prebuild_layout_result(prebuild_layout_json)
        pages = markup_source_pdf(source_pdf, source_pdf_markup_path, matching_results, prebuilt_layout_result, mark_symbols)
        for i, page in enumerate(pages):
            if page == 0:
                continue
            source_pages[i] = (source_pdf.stem, page)

    # Summarize in csv
    matching_log_csv = excel_path.parent / f"{excel_path.stem}_matching.csv"
    with matching_log_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        # ヘッダ
        writer.writerow(["cell", "value", "match", "source", "page_no" "reason", "symbol"])
        for inp, source_page, matching_result, symbol in zip(inputs, source_pages, matching_results, mark_symbols):
            writer.writerow([
                inp.cell,
                inp.value,
                matching_result.match,
                source_page[0] if matching_result.match else "-",
                source_page[1] if matching_result.match else "-",
                matching_result.reason,
                symbol if matching_result.match else "-"
            ])
