
from pathlib import Path
from openpyxl import load_workbook
from langchain_openai import ChatOpenAI

from excel_sheet_matching_agent.load_xlsx_llm import sheet2str, extract_inputs



def test_to_str():
    sheet_path = Path("examples/計算シートサンプル.xlsx")
    wb = load_workbook(sheet_path, data_only=False)
    print(sheet2str(wb["シート1"]))

def test_extract_inputs():
    sheet_path = Path("examples/計算シートサンプル.xlsx")
    wb = load_workbook(sheet_path, data_only=False)
    llm = ChatOpenAI(model="gpt-4o-mini", temperature=0)
    inputs = extract_inputs(llm, wb["シート1"])
    for input in inputs:
        print(input)
